"""
查询模块API视图 - 分页查询、筛选功能
"""
import json
import re
import datetime
import mysql.connector
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from connections.models import MySQLConnection
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi


def validate_identifier(name):
    """验证标识符（表名、列名）是否合法"""
    if not name or not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', name):
        return False
    return True


def parse_sort_param(sort_column, sort_order='asc'):
    """解析排序参数"""
    if not validate_identifier(sort_column):
        return None
    
    sort_order = sort_order.lower()
    if sort_order not in ['asc', 'desc']:
        sort_order = 'asc'
    
    return f"`{sort_column}` {sort_order.upper()}"


def build_where_clause(filters, filter_logic='AND'):
    """
    构建WHERE子句
    
    Args:
        filters: 筛选条件列表 [{"column": "age", "operator": ">=", "value": 18}, ...]
        filter_logic: 条件组合逻辑 AND/OR
    
    Returns:
        tuple: (where_clause, params)
    """
    if not filters:
        return "", []
    
    conditions = []
    params = []
    
    allowed_operators = {
        'eq': '=', 'neq': '!=', 'gt': '>', 'lt': '<', 'gte': '>=', 'lte': '<=',
        'contains': 'LIKE', 'startswith': 'LIKE', 'endswith': 'LIKE', 'in': 'IN'
    }
    
    for f in filters:
        column = f.get('column')
        operator = f.get('operator', 'eq').lower()
        value = f.get('value')
        
        # 验证字段名
        if not validate_identifier(column):
            continue
        
        # 验证操作符
        if operator not in allowed_operators:
            continue
        
        sql_operator = allowed_operators[operator]
        
        # 根据操作符构建条件
        if operator == 'contains':
            conditions.append(f"`{column}` {sql_operator} %s")
            params.append(f"%{value}%")
        elif operator == 'startswith':
            conditions.append(f"`{column}` {sql_operator} %s")
            params.append(f"%{value}")
        elif operator == 'endswith':
            conditions.append(f"`{column}` {sql_operator} %s")
            params.append(f"{value}%")
        elif operator == 'in':
            if isinstance(value, list):
                placeholders = ', '.join(['%s'] * len(value))
                conditions.append(f"`{column}` {sql_operator} ({placeholders})")
                params.extend(value)
        else:
            conditions.append(f"`{column}` {sql_operator} %s")
            params.append(value)
    
    if not conditions:
        return "", []
    
    filter_logic = filter_logic.upper()
    if filter_logic not in ['AND', 'OR']:
        filter_logic = 'AND'
    
    where_clause = f" WHERE {' ' + filter_logic + ' '.join(conditions)}"
    return where_clause, params


@login_required
@require_http_methods(["GET", "POST"])
def api_query_data(request):
    """
    通用数据查询API - 支持分页、排序、筛选
    
    GET /api/queries/data/?connection_id=1&table=users&page=1&page_size=20
    POST /api/queries/data/ (复杂筛选条件)
    
    请求参数:
        - connection_id: 数据库连接ID (必需)
        - table: 表名 (必需)
        - columns: 列名列表，逗号分隔，默认*
        - page: 页码，默认1
        - page_size: 每页大小，默认20，最大100
        - sort_column: 排序列
        - sort_order: 排序方向 asc/desc，默认asc
        - filters: JSON格式的筛选条件 (POST)
        - filter_logic: AND/OR，默认AND
    
    响应:
        {
            "code": 0,
            "message": "success",
            "data": {
                "list": [...],
                "pagination": {
                    "page": 1,
                    "page_size": 20,
                    "total": 100,
                    "total_pages": 5
                }
            }
        }
    """
    try:
        # 获取参数
        if request.method == "GET":
            connection_id = request.GET.get('connection_id')
            table = request.GET.get('table')
            columns = request.GET.get('columns', '*')
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 20))
            sort_column = request.GET.get('sort_column')
            sort_order = request.GET.get('sort_order', 'asc')
            filters = None
            filter_logic = request.GET.get('filter_logic', 'AND')
        else:  # POST
            data = json.loads(request.body)
            connection_id = data.get('connection_id')
            table = data.get('table')
            columns = data.get('columns', '*')
            page = int(data.get('page', 1))
            page_size = int(data.get('page_size', 20))
            sort_column = data.get('sort_column')
            sort_order = data.get('sort_order', 'asc')
            filters = data.get('filters')
            filter_logic = data.get('filter_logic', 'AND')
        
        # 验证必需参数
        if not connection_id or not table:
            return JsonResponse({
                "code": 400,
                "message": "缺少必需参数: connection_id 和 table"
            }, status=400)
        
        # 验证表名
        if not validate_identifier(table):
            return JsonResponse({
                "code": 400,
                "message": f"非法的表名: {table}"
            }, status=400)
        
        # 验证分页参数
        if page < 1:
            page = 1
        if page_size < 1:
            page_size = 20
        elif page_size > 100:
            page_size = 100  # 最大限制
        
        # 获取连接
        try:
            connection = MySQLConnection.objects.get(
                id=connection_id,
                created_by=request.user
            )
        except MySQLConnection.DoesNotExist:
            return JsonResponse({
                "code": 404,
                "message": "数据库连接不存在或无权限访问"
            }, status=404)
        
        # 构建查询
        connection_params = connection.get_connection_params()
        
        # 处理列
        if columns == '*':
            select_clause = '*'
        else:
            col_list = [c.strip() for c in columns.split(',')]
            # 验证所有列名
            for col in col_list:
                if not validate_identifier(col):
                    return JsonResponse({
                        "code": 400,
                        "message": f"非法的列名: {col}"
                    }, status=400)
            select_clause = ', '.join([f'`{c}`' for c in col_list])
        
        # 构建WHERE子句
        where_clause, query_params = build_where_clause(filters, filter_logic)
        
        # 构建ORDER BY子句
        order_clause = ""
        if sort_column:
            order_by = parse_sort_param(sort_column, sort_order)
            if order_by:
                order_clause = f" ORDER BY {order_by}"
        
        # 先查询总数
        count_sql = f"SELECT COUNT(*) as total FROM `{table}`{where_clause}"
        
        # 再查询数据
        offset = (page - 1) * page_size
        data_sql = f"SELECT {select_clause} FROM `{table}`{where_clause}{order_clause} LIMIT %s OFFSET %s"
        
        # 执行查询
        try:
            conn = mysql.connector.connect(**connection_params)
            cursor = conn.cursor(dictionary=True)
            
            # 查询总数
            if query_params:
                cursor.execute(count_sql, query_params)
            else:
                cursor.execute(count_sql)
            total = cursor.fetchone()['total']
            
            # 查询数据
            query_params_with_limit = query_params + [page_size, offset] if query_params else [page_size, offset]
            cursor.execute(data_sql, query_params_with_limit)
            data = cursor.fetchall()
            
            cursor.close()
            conn.close()
            
            # 计算总页数
            total_pages = (total + page_size - 1) // page_size if total > 0 else 0
            
            return JsonResponse({
                "code": 0,
                "message": "success",
                "data": {
                    "list": data,
                    "pagination": {
                        "page": page,
                        "page_size": page_size,
                        "total": total,
                        "total_pages": total_pages
                    }
                }
            })
            
        except mysql.connector.Error as e:
            return JsonResponse({
                "code": 500,
                "message": f"数据库查询错误: {str(e)}"
            }, status=500)
        
    except json.JSONDecodeError:
        return JsonResponse({
            "code": 400,
            "message": "JSON格式错误"
        }, status=400)
    except Exception as e:
        return JsonResponse({
            "code": 500,
            "message": f"服务器内部错误: {str(e)}"
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_get_table_structure(request):
    """
    获取表结构API

    GET /api/queries/table_structure/?connection_id=1&table=users&database=mydb

    请求参数:
        - connection_id: 数据库连接ID (必需)
        - table: 表名 (必需)
        - database: 数据库名 (可选，如果连接参数中没有指定数据库则必需)

    响应:
        {
            "code": 0,
            "message": "success",
            "data": [
                {
                    "Field": "id",
                    "Type": "int(11)",
                    "Null": "NO",
                    "Key": "PRI",
                    "Default": null,
                    "Extra": "auto_increment"
                },
                ...
            ]
        }
    """
    try:
        # 获取参数
        connection_id = request.GET.get('connection_id')
        table = request.GET.get('table')
        database = request.GET.get('database')

        # 验证必需参数
        if not connection_id or not table:
            return JsonResponse({
                "code": 400,
                "message": "缺少必需参数: connection_id 和 table"
            }, status=400)

        # 验证表名
        if not validate_identifier(table):
            return JsonResponse({
                "code": 400,
                "message": f"非法的表名: {table}"
            }, status=400)

        # 获取连接（所有登录用户都可以使用任意连接）
        try:
            connection = MySQLConnection.objects.get(id=connection_id)
        except MySQLConnection.DoesNotExist:
            return JsonResponse({
                "code": 404,
                "message": "连接不存在"
            }, status=404)

        # 执行查询获取表结构
        try:
            connection_params = connection.get_connection_params()

            # 如果指定了database参数，使用它
            if database:
                connection_params['database'] = database

            # 使用 mysql.connector 直接连接（不使用连接池，因为这是一个快速查询）
            import mysql.connector
            conn = mysql.connector.connect(**connection_params)
            cursor = conn.cursor(dictionary=True)

            # 执行 SHOW COLUMNS 查询获取表结构
            cursor.execute(f"SHOW COLUMNS FROM `{table}`")
            columns = cursor.fetchall()

            cursor.close()
            conn.close()

            return JsonResponse({
                "code": 0,
                "message": "success",
                "data": columns
            })

        except mysql.connector.Error as e:
            return JsonResponse({
                "code": 500,
                "message": f"数据库查询错误: {str(e)}"
            }, status=500)

    except Exception as e:
        return JsonResponse({
            "code": 500,
            "message": f"服务器内部错误: {str(e)}"
        }, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_export_excel(request):
    """
    导出查询结果为Excel文件API

    POST /api/queries/export_excel/

    请求体 (JSON):
    {
        "connection_id": 1,
        "database": "mydb",
        "sql": "SELECT * FROM users WHERE age > 18"
    }

    响应: Excel文件内容
    """
    try:
        # 解析请求体 - 支持JSON和form-data两种格式
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            connection_id = data.get('connection_id')
            database = data.get('database')
            sql = data.get('sql', '').strip()
        else:
            connection_id = request.POST.get('connection_id')
            database = request.POST.get('database')
            sql = request.POST.get('sql', '').strip()

        # 验证必需参数
        if not connection_id:
            return JsonResponse({
                "code": 400,
                "message": "缺少必需参数: connection_id"
            }, status=400)
        if not sql:
            return JsonResponse({
                "code": 400,
                "message": "缺少必需参数: sql"
            }, status=400)

        # 只允许 SELECT 查询
        sql_upper = sql.upper()
        if not sql_upper.startswith('SELECT'):
            return JsonResponse({
                "code": 400,
                "message": "只允许执行 SELECT 查询"
            }, status=400)

        # 检查危险关键字（简单防护）
        dangerous_keywords = ['DELETE', 'DROP', 'TRUNCATE', 'INSERT', 'UPDATE', 'ALTER']
        for keyword in dangerous_keywords:
            if keyword in sql_upper and keyword not in sql_upper.split('FROM')[0].split('WHERE')[0]:
                pattern = r'\b' + keyword + r'\b'
                if re.search(pattern, sql_upper):
                    return JsonResponse({
                        "code": 400,
                        "message": f"检测到危险关键字: {keyword}"
                    }, status=400)

        # 获取连接并检查权限
        try:
            if request.user.role == 'admin':
                connection = MySQLConnection.objects.get(id=connection_id)
            else:
                connection = MySQLConnection.objects.get(
                    id=connection_id,
                    created_by=request.user
                )
        except MySQLConnection.DoesNotExist:
            return JsonResponse({
                "code": 404,
                "message": "连接不存在或无权限访问"
            }, status=404)

        # 执行查询
        import time
        start_time = time.time()

        try:
            from connections.pool import get_connection_from_pool, release_connection

            connection_params = connection.get_connection_params()
            # 如果指定了database参数，需要先切换到该数据库
            if database:
                # 创建一个新的不包含数据库的连接参数，用于获取连接
                pool_params = connection_params.copy()
                pool_params.pop('database', None)

                conn = get_connection_from_pool(pool_params)
                cursor = conn.cursor(dictionary=True)

                # 先切换到指定数据库
                cursor.execute(f"USE `{database}`")

                # 再执行查询
                cursor.execute(sql)
                rows = cursor.fetchall()

                # 获取列名
                columns = [desc[0] for desc in cursor.description] if cursor.description else []

                cursor.close()
                release_connection(conn)
            else:
                # 没有指定数据库，使用连接池默认方式
                conn = get_connection_from_pool(connection_params)
                cursor = conn.cursor(dictionary=True)
                cursor.execute(sql)
                rows = cursor.fetchall()

                # 获取列名
                columns = [desc[0] for desc in cursor.description] if cursor.description else []

                cursor.close()
                release_connection(conn)

            # 应用脱敏规则
            from desensitization.utils import apply_masking_rules
            rows = apply_masking_rules(connection, sql, rows, request.user)

            # 生成Excel文件
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            import io

            wb = Workbook()
            ws = wb.active
            ws.title = '查询结果'

            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # 写入表头
            for col_num, column_title in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num, value=column_title)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                # 设置列宽
                column_letter = chr(64 + col_num) if col_num <= 26 else 'A' + chr(64 + col_num - 26)
                ws.column_dimensions[column_letter].width = max(15, len(str(column_title)) * 1.2)

            # 写入数据
            for row_num, row in enumerate(rows, 2):
                for col_num, column_title in enumerate(columns, 1):
                    cell_value = row.get(column_title, '')
                    cell = ws.cell(row=row_num, column=col_num, value=str(cell_value) if cell_value is not None else '')
                    # 设置数据列对齐
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # 构建响应
            from django.http import HttpResponse
            response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="query_result.xlsx"'
            return response

        except mysql.connector.Error as e:
            execution_time = (time.time() - start_time) * 1000
            return JsonResponse({
                "code": 500,
                "message": f"数据库查询错误: {str(e)}"
            }, status=500)

    except json.JSONDecodeError:
        return JsonResponse({
            "code": 400,
            "message": "JSON格式错误"
        }, status=400)
    except Exception as e:
        import traceback
        print(f"API导出Excel出错: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            "code": 500,
            "message": f"服务器内部错误: {str(e)}"
        }, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_execute_query(request):
    """
    执行 SQL 查询（仅支持 SELECT）

    POST /api/queries/execute/
    
    请求体:
    {
        "connection_id": 1,
        "sql": "SELECT * FROM users WHERE age > 18"
    }
    
    响应:
    {
        "code": 0,
        "message": "success",
        "data": {
            "columns": ["id", "name", "email", "age"],
            "rows": [...],
            "row_count": 2,
            "execution_time_ms": 15.23,
            "limited": false
        }
    }
    """
    try:
        # 解析请求体
        data = json.loads(request.body)
        connection_id = data.get('connection_id')
        sql = data.get('sql', '').strip()
        database = data.get('database', None)
        page = data.get('page', 1)
        page_size = data.get('page_size', 20)

        # 验证必需参数
        if not connection_id:
            return JsonResponse({
                "code": 400,
                "message": "缺少必需参数: connection_id"
            }, status=400)

        if not sql:
            return JsonResponse({
                "code": 400,
                "message": "缺少必需参数: sql"
            }, status=400)

        # 只允许 SELECT 查询
        sql_upper = sql.upper()
        if not sql_upper.startswith('SELECT'):
            return JsonResponse({
                "code": 400,
                "message": "只允许执行 SELECT 查询"
            }, status=400)

        # 检查危险关键字（简单防护）
        dangerous_keywords = ['DELETE', 'DROP', 'TRUNCATE', 'INSERT', 'UPDATE', 'ALTER']
        for keyword in dangerous_keywords:
            if keyword in sql_upper and keyword not in sql_upper.split('FROM')[0].split('WHERE')[0]:
                # 更严格的检查：确保关键字不是列名的一部分
                pattern = r'\b' + keyword + r'\b'
                if re.search(pattern, sql_upper):
                    return JsonResponse({
                        "code": 400,
                        "message": f"检测到危险关键字: {keyword}"
                    }, status=400)

        # 获取连接（所有登录用户都可以使用任意连接）
        try:
            connection = MySQLConnection.objects.get(id=connection_id)
        except MySQLConnection.DoesNotExist:
            return JsonResponse({
                "code": 404,
                "message": "连接不存在"
            }, status=404)

        # 执行查询
        import time
        start_time = time.time()

        try:
            from connections.pool import get_connection_from_pool, release_connection

            connection_params = connection.get_connection_params()

            # 如果指定了database参数，需要先切换到该数据库
            if database:
                # 创建一个新的不包含数据库的连接参数，用于获取连接
                pool_params = connection_params.copy()
                pool_params.pop('database', None)

                conn = get_connection_from_pool(pool_params)
                cursor = conn.cursor(dictionary=True)

                # 先切换到指定数据库
                cursor.execute(f"USE `{database}`")
            else:
                # 没有指定数据库，使用连接池默认方式
                conn = get_connection_from_pool(connection_params)
                cursor = conn.cursor(dictionary=True)

            # 执行查询 - 采用简单直接的方法
            sql_upper = sql.upper()

            if 'LIMIT' in sql_upper or 'OFFSET' in sql_upper:
                # 如果SQL已经包含LIMIT或OFFSET，直接执行
                cursor.execute(sql)
                rows = cursor.fetchall()
                total_count = len(rows)
                limited = False
            else:
                # 在内存中分页 - 简单且稳健
                cursor.execute(sql)
                all_rows = cursor.fetchall()
                total_count = len(all_rows)
                offset = (page - 1) * page_size
                rows = all_rows[offset:offset + page_size]
                limited = True

            # 获取列名
            columns = [desc[0] for desc in cursor.description] if cursor.description else []

            cursor.close()
            release_connection(conn)

            # 应用脱敏规则
            from desensitization.utils import apply_masking_rules
            rows = apply_masking_rules(connection, sql, rows, request.user)

            execution_time = (time.time() - start_time) * 1000

            # 记录查询历史和审计日志（不记录 sql，查询历史已有）
            try:
                from queries.models import QueryHistory
                QueryHistory.objects.create(
                    user=request.user,
                    connection=connection,
                    sql=sql,
                    execution_time=execution_time
                )
                # 审计日志不记录 sql
                from audit.utils import create_audit_log
                create_audit_log(
                    user=request.user,
                    action='query',
                    ip_address=request.META.get('REMOTE_ADDR'),
                    connection_id=connection.id,
                    execution_time=execution_time
                )
            except Exception:
                pass

            # 计算总页数
            total_pages = (total_count + page_size - 1) // page_size if total_count > 0 else 0

            # 转换不可序列化的数据类型
            for row in rows:
                for key, value in row.items():
                    if isinstance(value, (set, frozenset)):
                        row[key] = list(value)
                    elif isinstance(value, datetime):
                        row[key] = value.isoformat()

            return JsonResponse({
                "code": 0,
                "message": "success",
                "data": {
                    "columns": columns,
                    "rows": rows,
                    "row_count": len(rows),
                    "total_count": total_count,
                    "page": page,
                    "page_size": page_size,
                    "total_pages": total_pages,
                    "execution_time_ms": round(execution_time, 2),
                    "limited": limited
                }
            })
            
        except mysql.connector.Error as e:
            execution_time = (time.time() - start_time) * 1000
            return JsonResponse({
                "code": 500,
                "message": f"数据库查询错误: {str(e)}"
            }, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({
            "code": 400,
            "message": "JSON格式错误"
        }, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            "code": 500,
            "message": f"服务器内部错误: {str(e)}"
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_get_table_row_count(request):
    """
    获取表行数API

    GET /api/queries/table_row_count/?connection_id=1&table=users&database=mydb

    请求参数:
        - connection_id: 数据库连接ID (必需)
        - table: 表名 (必需)
        - database: 数据库名 (可选)

    响应:
        {
            "code": 0,
            "message": "success",
            "data": {
                "row_count": 1000
            }
        }
    """
    try:
        connection_id = request.GET.get('connection_id')
        table = request.GET.get('table')
        database = request.GET.get('database')

        if not connection_id or not table:
            return JsonResponse({
                "code": 400,
                "message": "缺少必需参数: connection_id 和 table"
            }, status=400)

        if not validate_identifier(table):
            return JsonResponse({
                "code": 400,
                "message": f"非法的表名: {table}"
            }, status=400)

        try:
            if request.user.role == 'admin':
                connection = MySQLConnection.objects.get(id=connection_id)
            else:
                connection = MySQLConnection.objects.get(
                    id=connection_id,
                    created_by=request.user
                )
        except MySQLConnection.DoesNotExist:
            return JsonResponse({
                "code": 404,
                "message": "连接不存在或无权限访问"
            }, status=404)

        try:
            connection_params = connection.get_connection_params()
            if database:
                connection_params['database'] = database

            import mysql.connector
            conn = mysql.connector.connect(**connection_params)
            cursor = conn.cursor(dictionary=True)

            cursor.execute(f"SELECT COUNT(*) as total FROM `{table}`")
            result = cursor.fetchone()
            row_count = result['total'] if result and 'total' in result else 0

            cursor.close()
            conn.close()

            return JsonResponse({
                "code": 0,
                "message": "success",
                "data": {
                    "row_count": row_count
                }
            })

        except mysql.connector.Error as e:
            return JsonResponse({
                "code": 500,
                "message": f"数据库查询错误: {str(e)}"
            }, status=500)

    except Exception as e:
        return JsonResponse({
            "code": 500,
            "message": f"服务器内部错误: {str(e)}"
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_get_configs(request):
    """
    获取系统配置API

    GET /api/queries/configs/

    响应:
        {
            "code": 0,
            "message": "success",
            "data": {
                "max_query_results": 10000,
                "enable_masking": True,
                "export_max_rows": 100000,
                "tables_per_page": 5,
                "sql_query_page_size": 20
            }
        }
    """
    try:
        from queries.models import SystemConfig

        configs = {
            "max_query_results": SystemConfig.get_int_value('max_query_results', 10000),
            "enable_masking": SystemConfig.get_value('enable_masking', 'true').lower() == 'true',
            "export_max_rows": SystemConfig.get_int_value('export_max_rows', 100000),
            "tables_per_page": SystemConfig.get_int_value('tables_per_page', 5),
            "sql_query_page_size": SystemConfig.get_int_value('sql_query_page_size', 20),
            "max_pagination_pages": SystemConfig.get_int_value('max_pagination_pages', 3),
            "sidebar_default_width": SystemConfig.get_int_value('sidebar_default_width', 250),
            "sidebar_min_width": SystemConfig.get_int_value('sidebar_min_width', 100),
            "sidebar_max_width": SystemConfig.get_int_value('sidebar_max_width', 600),
        }
        return JsonResponse({
            "code": 0,
            "message": "success",
            "data": configs
        })
    except Exception as e:
        return JsonResponse({
            "code": 500,
            "message": f"服务器内部错误: {str(e)}"
        }, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST", "GET", "DELETE", "PUT"])
def api_saved_queries(request):
    """
    保存的查询API - 列出、保存、更新、删除查询

    GET /api/queries/saved/ - 获取当前用户的所有保存查询
    POST /api/queries/saved/ - 保存新查询
    PUT /api/queries/saved/ - 更新已有查询
    DELETE /api/queries/saved/ - 批量删除查询
    """
    from queries.models import SavedQuery

    if request.method == "GET":
        # 获取保存的查询列表
        queries = SavedQuery.objects.filter(user=request.user).select_related('connection')
        data = []
        for q in queries:
            data.append({
                "id": q.id,
                "name": q.name,
                "sql": q.sql,
                "connection_id": q.connection_id,
                "connection_name": q.connection.name if q.connection else None,
                "database": q.database,
                "created_at": q.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                "updated_at": q.updated_at.strftime('%Y-%m-%d %H:%M:%S')
            })
        return JsonResponse({
            "code": 0,
            "message": "success",
            "data": data
        })

    elif request.method == "POST":
        # 保存新查询
        try:
            data = json.loads(request.body)
            name = data.get('name', '').strip()
            sql = data.get('sql', '').strip()
            connection_id = data.get('connection_id')
            database = data.get('database', '')

            if not name or not sql:
                return JsonResponse({
                    "code": 400,
                    "message": "名称和SQL都不能为空"
                }, status=400)

            # 获取连接对象
            connection = None
            if connection_id:
                try:
                    connection = MySQLConnection.objects.get(id=connection_id)
                except MySQLConnection.DoesNotExist:
                    pass

            # 创建保存的查询
            saved_query = SavedQuery.objects.create(
                user=request.user,
                name=name,
                sql=sql,
                connection=connection,
                database=database
            )

            return JsonResponse({
                "code": 0,
                "message": "保存成功",
                "data": {
                    "id": saved_query.id
                }
            })

        except json.JSONDecodeError:
            return JsonResponse({
                "code": 400,
                "message": "JSON格式错误"
            }, status=400)
        except Exception as e:
            return JsonResponse({
                "code": 500,
                "message": f"保存失败: {str(e)}"
            }, status=500)

    elif request.method == "PUT":
        # 更新已有查询
        try:
            data = json.loads(request.body)
            query_id = data.get('id')
            name = data.get('name', '').strip()
            sql = data.get('sql', '').strip()

            if not query_id:
                return JsonResponse({
                    "code": 400,
                    "message": "查询ID不能为空"
                }, status=400)

            if not name or not sql:
                return JsonResponse({
                    "code": 400,
                    "message": "名称和SQL都不能为空"
                }, status=400)

            # 查找并更新查询（只允许更新当前用户的）
            try:
                saved_query = SavedQuery.objects.get(id=query_id, user=request.user)
                saved_query.name = name
                saved_query.sql = sql
                saved_query.save()

                return JsonResponse({
                    "code": 0,
                    "message": "更新成功",
                    "data": {
                        "id": saved_query.id
                    }
                })
            except SavedQuery.DoesNotExist:
                return JsonResponse({
                    "code": 404,
                    "message": "查询不存在或无权限"
                }, status=404)

        except json.JSONDecodeError:
            return JsonResponse({
                "code": 400,
                "message": "JSON格式错误"
            }, status=400)
        except Exception as e:
            return JsonResponse({
                "code": 500,
                "message": f"更新失败: {str(e)}"
            }, status=500)

    elif request.method == "DELETE":
        # 批量删除查询
        try:
            data = json.loads(request.body)
            query_ids = data.get('ids', [])

            if not query_ids:
                return JsonResponse({
                    "code": 400,
                    "message": "请选择要删除的查询"
                }, status=400)

            # 删除查询（只删除当前用户的）
            deleted, _ = SavedQuery.objects.filter(
                user=request.user,
                id__in=query_ids
            ).delete()

            return JsonResponse({
                "code": 0,
                "message": f"成功删除 {deleted} 个查询",
                "data": {
                    "deleted_count": deleted
                }
            })

        except json.JSONDecodeError:
            return JsonResponse({
                "code": 400,
                "message": "JSON格式错误"
            }, status=400)
        except Exception as e:
            return JsonResponse({
                "code": 500,
                "message": f"删除失败: {str(e)}"
            }, status=500)

